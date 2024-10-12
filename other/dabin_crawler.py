"""
爬虫抓取大饼所需的进销存报表,并导出为xlsx表格
接受OCR识别请求地址,  http://localhost:8557/ocr
"""
import os
import abc
import copy
import openpyxl
import traceback
import collections
import pandas as pd
from dateutil.relativedelta import relativedelta
from typing import Dict, List, Tuple
from datetime import datetime
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
from medicine_utils import init_chrome, analyze_website, SPFJWeb, DruggcWeb, LYWeb, INCAWeb, CaptchaSocketServer


class Golbal:
    """全局数据"""

    def __init__(self) -> None:
        self.week_interval = None  # 间隔周数
        self.start_date = None  # 开始时间
        self.download_path = None  # 下载路径
        self.chrome_path = None  # 谷歌浏览器路径
        self.chromedriver_path = None  # 谷歌浏览器驱动路径
        self.websites_path = None  # 库存网查明细的文件路径
        self.database_path = None  # 脚本产品库的文件路径
        self.save_path = None  # 保存地址
        self.data = None  # 数据

    def set_data(self, path, week_interval):
        self.week_interval = week_interval
        self.download_path = path
        self.chrome_path = os.path.join(path, r"..\chromedriver_mac_arm64_114\chrome114\App\Chrome-bin\chrome.exe")
        self.chromedriver_path = os.path.join(path, r"..\chromedriver_mac_arm64_114\chromedriver.exe")
        self.websites_path = os.path.join(path, "福建商业明细表(福建)22.2.10-主席.xlsx")
        self.database_path = os.path.join(path, "产品库-傅镔滢.xlsx")
        self.save_path = os.path.join(path, "data.xlsx")
        now_date = datetime.now()
        if week_interval == 4:
            self.start_date = (now_date - relativedelta(months=3)).replace(day=1)
            self.data = collections.defaultdict(MonthData)
        else:
            self.start_date = now_date - relativedelta(days=now_date.weekday() + 7 * (week_interval - 1))
            self.data = collections.defaultdict(WeekData)
        self.start_date = self.start_date.strftime("%Y-%m-%d")

    def get_id(self, client_name, production_name):
        """根据客户名称及商品名称获取唯一ID"""
        return f"{client_name}@@{production_name}"

    def split_id(self, id: str):
        """根据ID分割出客户名称及商品名称"""
        return id.split("@@")


GOL = Golbal()


class BaseData:
    """基础数据类"""

    def __init__(self) -> None:
        self.code = ElementData("批号", "")

    def add_code(self, code):
        if code in self.code.value:
            return
        elif self.code.value == "":
            self.code.value = code
        else:
            self.code.value += f"、{code}"


class WeekData(BaseData):
    """每个客户周数据类"""

    def __init__(self) -> None:
        super().__init__()
        self.purchase = ElementData("本周进货", 0)
        self.sales = ElementData("本周销货", 0)
        self.inventory = ElementData("本周库存", 0)


class MonthData(BaseData):
    """每个客户的月数据类"""

    def __init__(self) -> None:
        super().__init__()
        self.purchase = ElementData("本月进货", 0)
        self.sales = ElementData("本月销货", 0)


class ElementData:
    """各个元素数据类"""

    def __init__(self, name, value) -> None:
        self.name = name
        self.value = value


class WebAbstract(abc.ABC):

    @abc.abstractmethod
    def login(self, *arg, **args):
        pass

    @abc.abstractmethod
    def export_data(self, client_name) -> dict:
        """获取数据"""
        pass


class SPFJWebCustom(SPFJWeb, WebAbstract):

    def export_data(self, client_name):
        web_data = self.purchase_sale_stock(GOL.start_date)
        for product_name, purchase, sale, inventory in web_data:
            id = GOL.get_id(client_name, product_name)
            data = GOL.data[id]
            data.purchase.value += purchase
            data.sales.value += sale
            if hasattr(data, "inventory"):
                data.inventory.value += inventory
        web_data = super().get_inventory(GOL.start_date)
        for product_name, _, code in web_data:
            id = GOL.get_id(client_name, product_name)
            data = GOL.data[id]
            data.add_code(code)


class LYWebCustom(LYWeb, WebAbstract):

    def export_data(self, client_name) -> dict:
        for product_name, _, code in self.get_inventory():
            id = GOL.get_id(client_name, product_name)
            data = GOL.data[id]
            data.add_code(code)
        web_data = self.purchase_sale_stock(GOL.start_date)
        for product_name, purchase, sale, inventory in web_data:
            id = GOL.get_id(client_name, product_name)
            data = GOL.data[id]
            data.purchase.value += purchase
            data.sales.value += sale
            if hasattr(data, "inventory"):
                data.inventory.value += inventory


class DruggcWebCustom(DruggcWeb, WebAbstract):

    def export_data(self, client_name):
        web_data = self.get_inventory()
        for product_name, inventory, code in web_data:
            id = GOL.get_id(client_name, product_name)
            data = GOL.data[id]
            data.add_code(code)
            if hasattr(data, "inventory"):
                data.inventory.value += inventory
        web_data = self.get_purchase(GOL.start_date)
        for product_name, amount in web_data:
            id = GOL.get_id(client_name, product_name)
            data = GOL.data[id]
            data.purchase.value += amount
        web_data = self.get_sales(GOL.start_date)
        for product_name, amount in web_data:
            id = GOL.get_id(client_name, product_name)
            data = GOL.data[id]
            data.sales.value += amount


class INCAWebCustom(INCAWeb, WebAbstract):

    def export_data(self, client_name):
        web_data = self.get_inventory()
        for product_name, inventory, code in web_data:
            id = GOL.get_id(client_name, product_name)
            data = GOL.data[id]
            data.add_code(code)
            if hasattr(data, "inventory"):
                data.inventory.value += inventory
        web_data = self.get_purchase(GOL.start_date)
        for product_name, amount in web_data:
            id = GOL.get_id(client_name, product_name)
            data = GOL.data[id]
            data.purchase.value += amount
        web_data = self.get_sales(GOL.start_date)
        for product_name, amount in web_data:
            id = GOL.get_id(client_name, product_name)
            data = GOL.data[id]
            data.sales.value += amount


class DataToExcel:
    """读取xlsx文件,并修改里面的数值"""

    def __init__(self) -> None:
        self.wb, self.ws = self.init_ws()

    def init_ws(self) -> Tuple[Workbook, Worksheet]:
        """初始化工作表"""
        wb = openpyxl.load_workbook(GOL.save_path)
        ws = wb.active
        print("删除颜色")
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, Cell) and not isinstance(cell, MergedCell):
                    raise Exception("表格存在异常")
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        return wb, ws

    def write_to_excel(self):
        # 读取产品数据库
        name2standard = {}
        db = pd.read_excel(GOL.database_path)
        for _, row in db.iterrows():
            client_name = row.iloc[0]
            client_product_name: str = row.iloc[1]
            standard_name = row.iloc[2]
            if pd.isna(client_product_name):
                continue
            client_product_name = client_product_name.replace(" ", "")
            name2standard[GOL.get_id(client_name, client_product_name)] = GOL.get_id(client_name, standard_name)
        # 将网站抓取的数据整理成标准数据
        save_data: Dict[str, dict] = {}
        for key, value in GOL.data.items():
            if key not in name2standard:
                print(f"[警告]未在产品库中找到,请及时添加:{key}")
                continue
            save_data[name2standard[key]] = value
        # 写入数据
        print("开始写入数据")
        now_date = datetime.now()
        titles = [value for row in self.ws.iter_rows(min_row=2, max_row=2, values_only=True) for value in row]
        for row in self.ws.iter_rows(min_row=3):
            date: Cell = row[titles.index("统计日期")]
            person: Cell = row[titles.index("负责人")]
            last_inventory: Cell = row[titles.index("上周库存" if GOL.week_interval != 4 else "上月库存")]
            now_inventory: Cell = row[titles.index("本周库存" if GOL.week_interval != 4 else "本月库存")]
            if date.value.strftime("%Y/%m/%d") == now_date.strftime("%Y/%m/%d"):
                continue
            if person.value != '傅镔滢':
                continue
            company: Cell = row[titles.index("经销商业公司名称")]
            product_name: Cell = row[titles.index("产品名称、规格")]
            id = GOL.get_id(company.value, product_name.value)
            if id not in save_data:
                print(f"[警告]未在网站中查询到该商品信息,{id}")
                continue
            date.value = now_date
            last_inventory.value = now_inventory.value
            for _, data in vars(save_data[id]).items():
                if not isinstance(data, ElementData):
                    raise Exception("类型异常")
                if data.name == "本周销货":
                    sell_index = f"{get_column_letter(titles.index('本周销货') + 1)}{date.row}"
                    remark: Cell = row[titles.index("备注")]
                    remark.value = f"=({sell_index}-{data.value})"
                else:
                    cell: Cell = row[titles.index(data.name)]
                    cell.value = data.value
        # 保存文件
        self.wb.save(GOL.save_path)


def read_breakpoint():
    """读取断点数据"""
    datas = pd.read_excel(GOL.save_path, header=1)
    datas = datas[datas["负责人"] == "傅镔滢"]
    today = datetime.today().date()
    datas = datas[datas["统计日期"].dt.date == today]
    client_names = set(datas["经销商业公司名称"].tolist())
    print(f"断点信息:{client_names}")
    return client_names


def crawler_general(datas: List[dict], url2class: Dict[str, WebAbstract]):
    """抓取的通用方法"""
    for data in datas:
        client_name = data.pop("client_name")
        website_url = data.pop("website_url")
        web_class = url2class[website_url]
        try:
            web_class.login(**data)
            web_class.export_data(client_name)
        except Exception:
            print("-" * 150)
            print(f"脚本运行出现异常, 出错的截至问题公司:{client_name}")
            print(traceback.format_exc())
            print("去除该客户的全部数据")
            data_key = copy.deepcopy(list(GOL.data.keys()))
            for key in data_key:
                if client_name not in key:
                    continue
                GOL.data.pop(key)
            print("-" * 150)
            return True
    return False


def crawler_websites_data(websites_by_code: List[dict], websites_no_code: List[dict]):
    """从网站上抓取数据，并写入全局变量"""
    print("抓取无验证码的网站数据")
    driver = init_chrome(GOL.chromedriver_path, GOL.download_path, chrome_path=GOL.chrome_path, is_proxy=False)
    url2class: Dict[str, WebAbstract] = {
        SPFJWeb.url: SPFJWebCustom(driver),
        INCAWeb.url: INCAWebCustom(driver, GOL.download_path),
        LYWeb.url: LYWebCustom(driver)
    }
    is_error = crawler_general(websites_no_code, url2class)
    if is_error:
        return
    print("关闭浏览器")
    driver.quit()
    print("抓取有验证码的网站数据")
    sock = CaptchaSocketServer()
    driver = init_chrome(GOL.chromedriver_path, GOL.download_path, chrome_path=GOL.chrome_path)
    url2class: Dict[str, WebAbstract] = {
        DruggcWeb.url: DruggcWebCustom(driver, sock, GOL.download_path)
    }
    crawler_general(websites_by_code, url2class)
    print("关闭浏览器")
    driver.quit()


def main(path, week):
    print("设置全局数据")
    GOL.set_data(path, week)
    print("读取断点数据")
    breakpoint_names = read_breakpoint()
    print("针对网站数据进行分类")
    websites_by_code, websites_no_code, _ = analyze_website(GOL.websites_path, breakpoint_names)
    print("从网站上爬取数据")
    crawler_websites_data(websites_by_code, websites_no_code)
    print("定义数据写入类")
    writer = DataToExcel()
    print("将数据写入到excel表格中")
    writer.write_to_excel()
    print("程序运行已完成")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("-p", "--path", type=str, default=r"E:\NewFolder\dabin", help="数据文件的所在文件夹地址")
    parser.add_argument("-w", "--week", type=int, default=4, help="间隔周数")
    opt = {key: value for key, value in parser.parse_args()._get_kwargs()}
    main(opt["path"], opt["week"])
