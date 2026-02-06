"""
爬虫抓取夏云所需的销售数据报表,并填入已有的数据表中
运行时候浏览器要全屏,否则有些按钮会没显示
"""
import os
import glob
import time
import shutil
import warnings
import calendar
import xlwings as xw
import pandas as pd
from copy import copy
from xlwings.main import Sheet
from dateutil.relativedelta import relativedelta
from decimal import Decimal, InvalidOperation
from abc import ABC, abstractmethod
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, JavascriptException
from typing import Dict, List
from datetime import timedelta, date, datetime
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, Font, NamedStyle, Side, Border
from openpyxl.utils.cell import get_column_letter, column_index_from_string
warnings.simplefilter("ignore")  # 忽略pandas使用openpyxl读取excel文件的警告
import matplotlib
matplotlib.use('Agg')  # 使用非 GUI 后端

class SavePath:
    """保存地址"""

    def __init__(self) -> None:
        self.operate_detail = None  # 营业明细表
        self.synthesize_operate = None  # 综合营业统计表
        self.synthesize_income = None  # 综合收款统计表
        self.store_consume = None  # 储值消费汇总表
        self.member_addition = None  # 会员新增情况统计表
        self.pay_settlement = None  # 支付结算表
        self.pay_detail = None  # 支付明细表

        self.take_out = None  # 外卖收入汇总表
        self.eleme_bill = None  # 饿了么账单明细
        self.autotrophy_meituan = None  # 自营外卖/自提订单明细表(美团)
        self.autotrophy_dada = None  # 自营外卖(达达)


class GolbalData:
    """全局数据"""

    def __init__(self) -> None:
        self.store_name = None  # 店名
        self.days, self.last_month = self.generate_last_month()
        self.save_path = SavePath()

    def generate_last_month(self):
        """获取上个月所有日期"""
        today = date.today()
        first_day_of_this_month = today.replace(day=1)
        last_day_of_last_month = first_day_of_this_month - timedelta(days=1)
        year, month = last_day_of_last_month.year, last_day_of_last_month.month
        num_days = calendar.monthrange(year, month)[1]
        last_months = [date(year, month, day).strftime("%y.%m.%d") for day in range(1, num_days + 1)]
        return last_months, last_day_of_last_month


class TradeMenu:
    """店名名单"""

    def __init__(self) -> None:
        self.huming = "Bread（湖明店）"
        self.ruijing = "Bread（瑞景店）"
        self.sweet = "Sweet"


GOL = GolbalData()
TM = TradeMenu()


class PerDayData:
    """营业明细统计表的每日数据类"""

    def __init__(self) -> None:
        self.cash = None  # 现金
        self.wechat = None  # 第三方收入（微信）
        self.eat_in = None  # 堂食扫码收入
        self.dianping = None  # 美团，大众点评
        self.ele_me = None  # 饿了么收入
        self.member_cash = []  # 会员充值现金收入
        self.member_wechat = []  # 会员充值第三方支付（微信）
        self.member_scan = []  # 会员扫码充值
        self.settlement_amount = None  # 实际结算金额
        self.main_consume = None  # 会员主账号消费
        self.gift_consume = None  # 会员赠送账号消费
        self.main_paid = None  # 会员主账号充值
        self.gift_paid = None  # 会员赠送账号充值
        self.new_member = None  # 新开卡会员
        self.ele_me_free = None  # 饿了么优免金额
        self.other_free = None  # 其他优免金额
        self.hand_charge = None  # 手续费
        self.pubilc_relation_paid = []  # 公关充值金额
        self.pubilc_relation_income = None  # 公关收入


class GetOperateDetail:
    """营业明细表的数据获取类"""

    def __init__(self, path) -> None:
        self.save_data: Dict[str, PerDayData] = self.init_save_data()
        self.wb = self.init_excel(path)
        self.ws = self.wb.active

    def init_excel(self, path):
        """调整营业明细表的模板"""
        days_len = len(GOL.days)
        if days_len == 31:
            return load_workbook(path)
        with xw.App(visible=False) as app:
            with app.books.open(path) as wb:
                ws = wb.sheets["营业月报"]
                for i in range(31 - days_len):
                    del_row = 35 - i
                    ws.range(f'{del_row}:{del_row}').api.EntireRow.Delete(Shift=-4162)
                wb.save(GOL.save_path.operate_detail)
        return load_workbook(GOL.save_path.operate_detail)

    def init_save_data(self) -> Dict[str, PerDayData]:
        """初始化每天的保存数据"""
        return {key: PerDayData() for key in GOL.days}

    def write_and_save(self):
        # 设置开始、结束时间
        self.ws.cell(2, 2, GOL.days[0])
        self.ws.cell(3, 2, GOL.days[-1])
        # 设置每天的数据
        for index, day in enumerate(GOL.days):
            row_index = 5 + index
            day_data = self.save_data[day]
            self.ws.cell(row_index, 2, day)
            self.ws.cell(row_index, 3, day_data.cash)
            self.ws.cell(row_index, 4, day_data.wechat)
            self.ws.cell(row_index, 5, day_data.eat_in)
            self.ws.cell(row_index, 6, day_data.dianping)
            self.ws.cell(row_index, 7, day_data.ele_me)
            self.ws.cell(row_index, 8, day_data.member_cash)
            self.ws.cell(row_index, 9, day_data.member_wechat)
            self.ws.cell(row_index, 10, day_data.member_scan)
            self.ws.cell(row_index, 13, day_data.settlement_amount)
            self.ws.cell(row_index, 15, day_data.main_consume)
            self.ws.cell(row_index, 16, day_data.gift_consume)
            self.ws.cell(row_index, 17, day_data.main_paid)
            self.ws.cell(row_index, 18, day_data.gift_paid)
            self.ws.cell(row_index, 19, day_data.new_member)
            self.ws.cell(row_index, 24, day_data.ele_me_free)
            self.ws.cell(row_index, 25, day_data.other_free)
            self.ws.cell(row_index, 31, day_data.hand_charge)
            self.ws.cell(row_index, 32, day_data.pubilc_relation_paid)
            self.ws.cell(row_index, 33, day_data.pubilc_relation_income)
        # 保存文件
        self.wb.save(GOL.save_path.operate_detail)

    def read_general_business(self):
        """读取综合营业统计表的相关数据"""
        data = pd.read_excel(GOL.save_path.synthesize_operate, header=None)
        row2 = data.iloc[2].apply(replace_parentheses)
        row3 = data.iloc[3].apply(replace_parentheses)
        row4 = data.iloc[4].apply(replace_parentheses)
        ele_me_i = get_3row_index(row2, row3, row4, "渠道营业构成", "饿了么外卖", "营业收入（元）", is_must=False)
        dianping_wx_i = get_3row_index(row2, row3, row4, "营业收入构成", "美团/大众点评支付", "微信", is_must=False)
        dianping_mt_i = get_3row_index(row2, row3, row4, "营业收入构成", "美团/大众点评支付", "美团支付", is_must=False)
        cach_i = get_3row_index(row2, row3, row4, "营业收入构成", "现金", "人民币", is_must=False)
        eat_in_i_list = list(range(*get_3row_index(row2, row3, row4, "营业收入构成", "扫码支付", None)))
        pubilc_relation_income_i = get_3row_index(
            row2, row3, row4, "营业收入构成", "自定义记账", "公关/奖品/活动/无实质性收入（自）", is_must=False)
        wechat_i = get_3row_index(row2, row3, row4, "营业收入构成", "自定义记账", [
                                  "微信收款（店长号收款）（自）", "微信店长号收款（自）", "店长微信收款收入（自）"])
        ele_me_free_i = get_3row_index(row2, row3, row4, "支付优惠构成", "外卖", "饿了么外卖")  # 56
        other_free_i = get_3row_index(row2, row3, row4, "折扣优惠构成", "小计", "小计")  # 68
        for row in data.iloc[5:].itertuples():
            row = row[1:]
            day_str = row[0]
            if day_str == "合计":
                break
            day_str: str = day_str[2:]
            day_str = day_str.replace("/", ".")
            day_data = self.save_data[day_str]
            day_data.cash = row[cach_i] if cach_i is not None else 0
            day_data.wechat = row[wechat_i]
            day_data.eat_in = sum(list_generate(eat_in_i_list, row))
            day_data.ele_me = row[ele_me_i] if ele_me_i is not None else 0 
            day_data.dianping = sum([row[i] if i is not None else 0 for i in [dianping_wx_i, dianping_mt_i]])
            day_data.ele_me_free = row[ele_me_free_i]
            day_data.other_free = row[other_free_i]
            day_data.pubilc_relation_income = row[pubilc_relation_income_i] if pubilc_relation_income_i is not None else 0

    def read_general_collection(self):
        """读取综合收款统计表的相关数据"""
        data = pd.read_excel(GOL.save_path.synthesize_income, header=None)
        row2 = data.iloc[2].apply(replace_parentheses)
        row3 = data.iloc[3].apply(replace_parentheses)
        cash_i = get_2row_index(row2, row3, "现金", "人民币", is_must=False)
        scan_i_list = list(range(*get_row_range(row2, "扫码支付")))
        income_i = get_2row_index(row2, row3, "自定义记账", "公关/奖品/活动/无实质性收入（自）", is_must=False)
        wechat_i = get_2row_index(row2, row3, "自定义记账", ["微信收款（店长号收款）（自）", "微信店长号收款（自）", "店长微信收款收入（自）"])
        if wechat_i is None:
            raise Exception("找不到数据:微信收款（店长号收款）（自）, 微信店长号收款（自）, 店长微信收款收入（自）")
        for row in data.iloc[5:].itertuples():
            row = row[1:]
            day_str = row[0]
            if day_str == "合计":
                break
            day_str: str = day_str[2:]
            day_str = day_str.replace("/", ".")
            day_data = self.save_data[day_str]
            if row[1] != "会员充值":
                continue
            if row[2] in ["充值", "撤销充值"]:
                cash = row[cash_i] if cash_i is not None else 0
                wechat = row[wechat_i]
                scan = sum(list_generate(scan_i_list, row))
                income = 0 if income_i is None else row[income_i]
            elif row[2] == "退卡":
                cash = row[cash_i] if cash_i is not None else 0
                wechat = row[wechat_i]
                income = row[income_i] if income_i is not None else 0
                scan = sum(list_generate(scan_i_list, row))
                assert cash <= 0, "退卡金额应该小于0"
                assert wechat <= 0, "退卡金额应该小于0"
                assert income <= 0, "退卡金额应该小于0"
                assert scan <= 0, "退卡金额应该小于0"
            else:
                continue
            day_data.member_cash.append(cash)
            day_data.member_wechat.append(wechat)
            day_data.member_scan.append(scan)
            day_data.pubilc_relation_paid.append(income)
        # 整理数据
        for _, day_data in self.save_data.items():
            assert len(day_data.member_cash) <= 3, "会员充值、退卡数据各自最多只有一条"
            day_data.member_cash = sum(day_data.member_cash)
            assert len(day_data.member_wechat) <= 3, "会员充值、退卡数据各自最多只有一条"
            day_data.member_wechat = sum(day_data.member_wechat)
            assert len(day_data.member_scan) <= 3, "会员充值、退卡数据各自最多只有一条"
            day_data.member_scan = sum(day_data.member_scan)
            assert len(day_data.pubilc_relation_paid) <= 3, "公关/奖品/活动/无实质性收入的充值、退卡数据各自最多只有一条"
            day_data.pubilc_relation_paid = sum(day_data.pubilc_relation_paid)

    def read_store_consume(self):
        """读取储值消费汇总表的相关数据"""
        data = pd.read_excel(GOL.save_path.store_consume, header=None)
        row2 = data.iloc[2].apply(replace_parentheses)
        row3 = data.iloc[3].apply(replace_parentheses)
        main_consume_i = get_2row_index(row2, row3, "净储值消费金额", "本金（元）")
        gift_consume_i = get_2row_index(row2, row3, "净储值消费金额", "赠金（元）")
        main_paid_i = get_2row_index(row2, row3, "净储值金额", "本金（元）")
        gift_paid_i = get_2row_index(row2, row3, "净储值金额", "赠金（元）")
        for row in data.iloc[4:].itertuples():
            row = row[1:]  # 去除索引
            day_str = row[0]
            if day_str == "合计":
                break
            day_str: str = day_str[2:]
            day_str = day_str.replace("/", ".")
            day_data = self.save_data[day_str]
            day_data.main_consume = row[main_consume_i]  # 会员主账号消费
            day_data.gift_consume = row[gift_consume_i]  # 会员赠送账号消费
            day_data.main_paid = row[main_paid_i]  # 会员主账号充值
            day_data.gift_paid = row[gift_paid_i]  # 会员赠送账号充值


    def read_newly_increased(self):
        """"读取会员新增情况统计表的相关数据"""
        data = pd.read_excel(GOL.save_path.member_addition, header=None)
        assert data.iloc[2, 0] == "日期", "表格发生变化，请联系管理员"
        assert data.iloc[2, 1] == "合计", "表格发生变化，请联系管理员"
        for row in data.iloc[3:].itertuples():
            row = row[1:]
            if row[0] == "合计":
                break
            day_str: str = row[0][2:]
            day_str = day_str.replace("-", ".")
            day_data = self.save_data[day_str]
            day_data.new_member = row[1]

    def read_pay_settlement(self):
        """读取支付结算表的相关数据"""
        data = pd.read_excel(GOL.save_path.pay_settlement, header=None)
        row2 = data.iloc[2].apply(replace_parentheses)
        assert row2[2] == "结算日期", "表格发生变化，请联系管理员"
        assert row2[3] == "交易金额（元）", "表格发生变化，请联系管理员"
        assert row2[5] == "手续费（元）", "表格发生变化，请联系管理员"
        for row in data.iloc[3:len(GOL.days) + 3].itertuples():
            row = row[1:]
            day_str = row[2]
            if pd.isna(day_str):
                break
            day_str: str = day_str[2:]
            day_str = day_str.replace("-", ".")
            day_data = self.save_data[day_str]
            day_data.hand_charge = row[5]
            day_data.settlement_amount = row[3]


class WebCrawler(ABC):

    def __init__(self, driver: Chrome, download_path) -> None:
        self._download_timeout = 60
        self._download_path = download_path
        self._name2save = self._init_name2save()
        self._driver = driver
        self._action = ActionChains(self._driver)

    @abstractmethod
    def _init_name2save(self) -> dict:
        """初始化下载文件与文件保存地址的对应关系"""

    def wait_download(self, name, name_key=None):
        """等待下载

        Args:
            name (str): 下载的文件名称
            name_key (str, optional): 对应保存地址获取的键值. Defaults to None.
        """
        print(f"等待下载文件:{name}")
        name_key = name if name_key is None else name_key
        file_path: str = ""
        st = time.time()
        while True:
            if (time.time() - st) > self._download_timeout:
                raise Exception(f"Waiting download timeout:{name}")
            download_files = glob.glob(os.path.join(self._download_path, name))
            if len(download_files) == 0:
                time.sleep(0.5)
                continue
            file_path = download_files[0]
            break
        print(f"文件已经在下载:{os.path.basename(file_path)}")
        if not file_path.endswith(".xlsx"):
            file_path = file_path.replace(".crdownload", "")
            st = time.time()
            while True:
                if (time.time() - st) > self._download_timeout:
                    raise Exception("Waiting download finnish timeout.")
                if os.path.exists(file_path):
                    break
                time.sleep(0.5)
        time.sleep(3)
        # 移动下载文件
        shutil.move(file_path, self._name2save[name_key])
        print(f"文件已移动到相应位置:{self._name2save[name_key]}")

class DadaCrawler(WebCrawler):

    def _init_name2save(self):
        return {
            "门店订单明细": GOL.save_path.autotrophy_dada
        }

    def login(self):
        """登录达达网站"""
        self._driver.get(r"https://newopen.imdada.cn/#/manager/shop/report/order")
        time.sleep(5)

    def download_store_report(self):
        """下载门店报表"""
        if os.path.exists(self._name2save["门店订单明细"]):
            print("门店订单明细已下载,不重复下载")
            return
        # 日期选择
        pattern = (By.XPATH, ".//input[@placeholder='请选择时间']")
        WebDriverWait(self._driver, 10).until(EC.element_to_be_clickable(pattern))
        start_select_ele, end_select_ele = self._driver.find_elements(*pattern)
        start_ele, end_ele = self._driver.find_elements(By.CLASS_NAME, "datepicker")
        self.__date_selection(start_select_ele, start_ele, 1)
        self.__date_selection(end_select_ele, end_ele, GOL.days[-1].split(".")[-1])
        self._driver.find_element(By.XPATH, "//span[text()='搜索']/..").click()
        load_icon = self._driver.find_element(By.CLASS_NAME, "loading-mask")
        try:
            WebDriverWait(self._driver, 10).until(EC.visibility_of(load_icon))
        except TimeoutException:
            print("订单报表查询的加载图标已提前完成")
        time.sleep(1)
        WebDriverWait(self._driver, 120).until(EC.invisibility_of_element(load_icon))
        # 申请报表下载
        if GOL.store_name == TM.huming:
            btn_str = "//div[text()='Still Bread（湖明店）']/../..//span[text()='下载门店订单明细']"
        elif GOL.store_name == TM.ruijing:
            btn_str = "//div[text()='Still bread 还是面包厨房（华瑞花园1期店）']/../..//span[text()='下载门店订单明细']"
        elif GOL.store_name == TM.sweet:
            btn_str = "//div[text()='Still sweet']/../..//span[text()='下载门店订单明细']"
        else:
            raise Exception("未知店名")
        WebDriverWait(self._driver, 20).until(EC.presence_of_element_located((By.XPATH, btn_str)))
        ele = WebDriverWait(self._driver, 20).until(EC.element_to_be_clickable((By.XPATH, btn_str)))
        ele.click()
        time.sleep(3)
        WebDriverWait(self._driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[@class='modal-header']//h5[text()='报表下载提示']")))
        ele = WebDriverWait(self._driver, 60).until(EC.element_to_be_clickable((By.XPATH, "//div[@class='modal-body']//a[text()='下载列表']")))
        ele.click()
        time.sleep(5)
        # 跳转到下载页面
        WebDriverWait(self._driver, 60).until(EC.presence_of_element_located((By.XPATH, "//div[text()=' 订单报表']/..")))
        ele = WebDriverWait(self._driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//div[text()=' 订单报表']/..")))
        if "active" not in ele.get_attribute("class"):
            print("展开订单列表")
            ele.click()
        WebDriverWait(ele, 30).until(EC.presence_of_element_located((By.XPATH, ".//a[text()='下载列表']")))
        ele = WebDriverWait(ele, 10).until(EC.element_to_be_clickable((By.XPATH, ".//a[text()='下载列表']")))
        for i in range(5):
            if "active" in ele.get_attribute("class"):
                break
            ele.click()
            time.sleep(1)
            print("点击下载列表出现问题，重新点击")
        else:
            raise Exception("点击下载列表错误")
        try:
            WebDriverWait(self._driver, 10).until(EC.visibility_of(load_icon))
        except TimeoutException:
            print("下载列表的加载图标已提前完成")
        WebDriverWait(self._driver, 60).until(EC.invisibility_of_element(load_icon))
        for i in range(5):
            download_result = self.__download_file()
            print(f"第{i}次的文件下载结果:{download_result}")
            if download_result:
                break

    def __download_file(self):
        # 定位到所需要下载的那一行
        WebDriverWait(self._driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[text()='申请日期']")))
        now_str = date.today().strftime("%Y-%m-%d")
        true_content = f"(20{GOL.days[0].replace('.', '-')} ~ 20{GOL.days[-1].replace('.', '-')})"
        tr_ele_list = self._driver.find_elements(By.XPATH, f"//div[text()='{now_str}']/../..")
        for tr_ele in tr_ele_list:
            td_eles = tr_ele.find_elements(By.TAG_NAME, "td")
            content_ele = td_eles[2].find_elements(By.XPATH, "./div/div/div")[1]
            if content_ele.text != true_content:
                continue
            content_status = td_eles[3].find_element(By.XPATH, "./div/div/div")
            if content_status.text != "已生成":
                print("下载报表未生成完毕,等待1分钟后刷新")
                time.sleep(60)
                self._driver.refresh()
                time.sleep(3)
                return False
            break
        else:
            raise Exception("找不到对应的下载信息")
        # 下载文件
        button = WebDriverWait(tr_ele, 30).until(EC.element_to_be_clickable((By.XPATH, ".//a[text()='下载']")))
        button.click()
        time.sleep(5)
        name = os.path.basename(button.get_attribute("href"))
        self.wait_download(name, "门店订单明细")
        return True

    def __date_selection(self, select_ele: WebElement, element: WebElement, value):
        """日期选择"""
        select_ele.click()
        WebDriverWait(self._driver, 1).until(EC.visibility_of(element))
        now_date_ele = element.find_element(By.CLASS_NAME, "datepicker-caption")
        need_date_str = GOL.last_month.strftime("%Y年%m月")
        while True:
            cur_date_str = now_date_ele.text
            if cur_date_str == need_date_str:
                break
            now_date = datetime.strptime(cur_date_str, "%Y年%m月").date()
            if now_date - GOL.last_month > timedelta(0):
                element.find_element(By.CLASS_NAME, "dada-ico-angle-left").click()
            else:
                element.find_element(By.CLASS_NAME, "dada-ico-angle-right").click()
            WebDriverWait(now_date_ele, 1).until(lambda ele: ele.text != cur_date_str)
        # 网页中的标签情况：<div class="datepicker-item-text active notinmonth">  
        ele = WebDriverWait(element, 10).until(EC.presence_of_element_located((By.XPATH, f".//div[not(contains(@class, 'notinmonth')) and text()='{value}']")))
        ele.click()


class MeiTuanCrawler(WebCrawler):
    """美团网站抓取类"""

    def _init_name2save(self):
        return {
            "综合营业统计": GOL.save_path.synthesize_operate,
            "综合收款统计": GOL.save_path.synthesize_income,
            "储值消费汇总表": GOL.save_path.store_consume,
            "会员新增情况统计表": GOL.save_path.member_addition,
            "支付结算": GOL.save_path.pay_settlement,
            "支付明细": GOL.save_path.pay_detail,
            "自营外卖_自提订单明细": GOL.save_path.autotrophy_meituan,
        }

    def login(self):
        """登录显示非法请求，不让操作"""
        # driver.get(r"https://pos.meituan.com")
        # wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "login-section-36Fr9")))
        # driver.switch_to.frame(driver.find_element(By.XPATH, "//iframe[@title='rms epassport']"))
        # wait.until(EC.visibility_of_element_located((By.XPATH, "//a[text()='帐号登录']")))
        # driver.find_element(By.XPATH, "//a[text()='帐号登录']").click()
        # wait.until(EC.visibility_of_element_located((By.ID, "password")))
        # clear_and_send(driver.find_element(By.ID, "account"), r"13194080718")
        # clear_and_send(driver.find_element(By.ID, "password"), r"stillrelax0601")
        # driver.find_element(By.XPATH, "//button[text()='登录']").click()
        self._driver.get(r"https://pos.meituan.com/web/operation/main#/")
        WebDriverWait(self._driver, 30).until(EC.title_is("美团管家"))
        time.sleep(3)

    def toggle_store(self, name):
        """切换店铺"""
        if name == TM.huming:
            ele_name = "Still bread KIT(湖明店）"
            pattern = (By.XPATH, f"//tbody[@class='saas-table-tbody']//td[text()='{ele_name}']")
        elif name == TM.ruijing:
            ele_name = "Still bread Kit·还是面包厨房（瑞景店）"
            pattern = (By.XPATH, f"//tbody[@class='saas-table-tbody']//td[text()='{ele_name}']")
        elif name == TM.sweet:
            ele_name = "Still sweet·还是甜品厨房（洪文店）"
            pattern = (By.XPATH, f"//tbody[@class='saas-table-tbody']//td[text()='{ele_name}']")
        else:
            raise Exception("店铺名称错误")
        btn = WebDriverWait(self._driver, 60).until(EC.element_to_be_clickable((By.CLASS_NAME, "perspective-switch")))
        if ele_name == btn.get_attribute("title"):
            print("当前所在店铺即为所需,不用切换")
            return
        for _ in range(6):
            btn.click()
            try:
                WebDriverWait(self._driver, 5).until(EC.element_to_be_clickable(pattern))
            except TimeoutException:
                continue
            break
        else:
            raise Exception("切换店铺失败,始终找不到弹窗")
        btn = self._driver.find_element(*pattern)
        btn.click()
        WebDriverWait(self._driver, 60).until(EC.invisibility_of_element(pattern))
        WebDriverWait(self._driver, 60).until(EC.element_to_be_clickable((By.CLASS_NAME, "perspective-switch")))
        print(f"成功切换到店铺:{name}")

    def download_synthesize_operate(self):
        """下载综合营业统计表"""
        menu_name, name = "营业报表", "综合营业统计"
        if os.path.exists(self._name2save[name]):
            print(f"{name}已存在，不再重新下载")
            return
        module = self._enter_main_module("报表中心")
        self._enter_rc_module(module, menu_name, name)
        # 选择日期
        self._wait_shadow_dom()
        action = self._click_normal_by_js("button.saas-btn-link span.saasicon-double-down")
        signal = self._wait_by_js("button.saas-btn-link span.saasicon-double-up")
        self._js_click(action, signal)  # 展开筛选
        self._js_click_start_date()
        self._js_click_last_month()
        self._js_click_search()
        self._clear_excel(name)  # 清理excel文件
        self._js_click_download()
        self.wait_download(f"*{name}*", name)  # 等待文件下载完成
        # 点击提醒
        # action = f"{self._js_span_find('我知道了')}.click()"
        # signal = f"!{self._js_shadow_root()}.querySelector('div[role=\"dialog\"]')"
        # self._js_click(action, self._js_execute_result(signal))

    def download_autotrophy(self):
        """下载自营外卖/自提订单明细表"""
        menu_name, name = "营业报表", "自营外卖/自提订单明细"
        download_name = name.replace("/", "_")
        if os.path.exists(self._name2save[download_name]):
            print(f"{download_name}已存在，不再重新下载")
            return
        module = self._enter_main_module("报表中心")
        submodule = self._enter_rc_module(module, menu_name, name)
        self._date_select_3(submodule)
        self._search(submodule)
        self._clear_excel(download_name)
        self._wait_search_finnsh_1(submodule)
        self._download_autotrophy_detail(submodule, download_name)

    def download_synthesize_income(self):
        """下载综合收款统计表"""
        menu_name, name = "收款报表", "综合收款统计"
        if os.path.exists(self._name2save[name]):
            print(f"{name}已存在，不再重新下载")
            return
        module = self._enter_main_module("报表中心")
        self._enter_rc_module(module, menu_name, name)
        self._wait_shadow_dom()  # 等待shadowRoot显现
        self._js_click_start_date()
        self._js_click_last_month()
        action = f"{self._js_span_find('按日')}.click() \n return true"
        signal = f"{self._js_span_find('按日')}.parentNode.className.includes('saas-radio-button-wrapper-checked')"
        self._js_click(action, self._js_execute_result(signal))
        action = f"{self._js_span_find('业务小类')}.click() \n return true"
        signal = f"{self._js_span_find('业务小类')}.parentNode.className.includes('saas-radio-button-wrapper-checked')"
        self._js_click(action, self._js_execute_result(signal))
        self._js_click_search()
        self._clear_excel(name)
        self._js_click_download()
        self.wait_download(f"*{name}*", name)

    def download_pay_settlement(self):
        """下载支付结算表"""
        menu_name, name = "收款报表", "支付结算"
        if os.path.exists(self._name2save[name]):
            print(f"{name}已存在，不再重新下载")
            return
        module = self._enter_main_module("报表中心")
        submodule = self._enter_rc_module(module, menu_name, name)
        self._date_select_1(submodule)
        self._pay_settlement_condition(submodule)
        self._search(submodule)
        self._clear_excel(name)
        self._wait_search_finnsh_1(submodule)
        self._download_direct(submodule, name, name)

    def download_pay_detail(self):
        """下载支付明细表"""
        menu_name, name = "收款报表", "支付明细"
        if os.path.exists(self._name2save[name]):
            print(f"{name}已存在，不再重新下载")
            return
        module = self._enter_main_module("报表中心")
        submodule = self._enter_rc_module(module, menu_name, name)
        self._date_select_1(submodule)
        self._search(submodule)
        self._clear_excel(name)
        self._wait_search_finnsh_1(submodule)
        self._download_direct(submodule, name, name)


    def download_store_consume(self):
        """下载储值消费汇总表"""
        menu_name = "数据报表"
        name = "储值余额变动汇总表（原储值消费汇总表）"
        iframe_url = '/web/crm-smart/report/dpaas-summary-store'
        download_name = "储值余额变动汇总表"
        save_name = "储值消费汇总表"
        if os.path.exists(self._name2save[save_name]):
            print(f"{save_name}已存在，不再重新下载")
            return
        module = self._enter_main_module("营销中心")
        self._enter_mc_module(module, menu_name, name, iframe_url)
        # 选择日期
        now_date = datetime.today()
        end_date = now_date.replace(day=1) - relativedelta(days=1)
        start_date = end_date.replace(day=1)
        self._wait_shadow_dom()
        action = self._click_normal_by_js('input[placeholder="开始日期"]', is_parent=True)
        signal = self._wait_by_js("div.saas-picker-dropdown-range", key_value="hidden", is_has="false")
        self._js_click(action, signal)  # 点击开始日期
        action = self._click_normal_by_js("button.saas-picker-header-prev-btn")
        signal = self._wait_by_js("div.saas-picker-dropdown-range button", text=f"{start_date.month}月")
        self._js_click(action, signal)  # 点击上个月按钮
        action = self._click_normal_by_js(f'td.saas-picker-cell[title="{start_date.strftime("%Y-%m-%d")}"]')
        signal = self._wait_by_js(f'td.saas-picker-cell[title="{start_date.strftime("%Y-%m-%d")}"]', key_value="selected")
        self._js_click(action, signal)  # 点击起始日期
        action = self._click_normal_by_js("li.saas-picker-ok button.saas-btn-primary")
        signal = self._wait_by_js("div.saas-picker-dropdown-range button", text=f"{now_date.month}月")
        self._js_click(action, signal)  # 点击确定
        action = self._click_normal_by_js("button.saas-picker-header-prev-btn")
        signal = self._wait_by_js("div.saas-picker-dropdown-range button", text=f"{start_date.month}月")
        self._js_click(action, signal)  # 点击上个月按钮
        action = self._click_normal_by_js(f'td.saas-picker-cell[title="{end_date.strftime("%Y-%m-%d")}"]')
        signal = self._wait_by_js(f'td.saas-picker-cell[title="{end_date.strftime("%Y-%m-%d")}"]', key_value="selected")
        self._js_click(action, signal)  # 点击结束日期
        action = self._click_normal_by_js("li.saas-picker-ok button.saas-btn-primary")
        signal = self._wait_by_js("div.saas-picker-dropdown-range", key_value="hidden", is_has="true")
        self._js_click(action, signal)  # 点击确定
        # 开始查询数据
        action = self._click_normal_by_js("div.saas-radio-group-outline span", text="日期")
        signal = self._wait_by_js("div.saas-radio-group-outline span", text="日期", key_value="checked", is_parent=True)
        self._js_click(action, signal)  # 点击按照日期统计
        action = self._click_normal_by_js("button.saas-btn-primary span.saasicon-search")
        signal = self._wait_search_btn_by_js(is_finish=False)
        self._js_click(action, signal)  # 点击查询
        signal = self._wait_search_btn_by_js()
        self._js_wait(signal)  # 等待查询完成
        # 开始导出文件
        self._clear_excel(name)
        action = self._click_normal_by_js("button.saas-btn-text span.saasicon-download")
        signal = self._wait_export("button.saas-btn-text span")
        self._js_click(action, signal)  # 点击导出
        self.wait_download(f"*{download_name}*", save_name)  # 等待导出完成

    def download_member_addition(self):
        """下载会员新增情况统计表"""
        menu_name = "数据报表"
        name = "会员新增分析"
        iframe_url = '/web/member/statistic/member-increase#/'
        save_name = "会员新增情况统计表"
        if os.path.exists(self._name2save[save_name]):
            print(f"{save_name}已存在，不再重新下载")
            return
        # 进入模块
        module = self._enter_main_module("营销中心")
        submodule = self._enter_mc_module(module, menu_name, name, iframe_url)
        self._toggle_old()
        self._date_select_3(submodule)
        self._search(submodule)
        self._clear_excel(name)
        self._wait_search_finnsh_1(submodule)
        self._download_direct(submodule, name, save_name)

    def _enter_main_module(self, name):
        """进入主模块的方法：运营中心、营销中心、库存管理、报表中心"""
        self._driver.switch_to.default_content()
        # 获取当前容器显示的模块内容
        pattern = (By.XPATH, "//div[@class='main-app']/div[@style='display: block;'][./*]")
        WebDriverWait(self._driver, 10).until(EC.presence_of_all_elements_located(pattern))
        current_module = self._driver.find_element(*pattern)
        # 进入所定模块
        pattern = (By.XPATH, f"//header//span[text()='{name}']/..")
        WebDriverWait(self._driver, 10).until(EC.visibility_of_element_located(pattern))
        element = self._driver.find_element(*pattern)
        if "active-first-menu" in element.get_attribute("class"):
            print(f"已在{name}模块")
            time.sleep(5)
            return current_module
        element.click()
        # 等待原有模块消失，即新模块显示
        WebDriverWait(current_module, 10).until(lambda ele: ele.get_attribute("style") == "display: none;")
        # 获取新模块
        pattern = (By.XPATH, "//div[@class='main-app']/div[@style='display: block;'][./*]")
        current_module = self._driver.find_element(*pattern)
        try:
            current_module.find_element(By.XPATH, f".//div[@role='tablist']//span[text()='{name}首页']")
        except NoSuchElementException as exc:
            raise Exception("进入模块出现问题") from exc
        time.sleep(5)
        return current_module

    def _enter_rc_module(self, module, menu_name, name):
        """进入报表中心的子模块"""
        print(f"进入报表中心-{menu_name}-{name}")
        submodule = self._get_statement_submodule(module)
        if self._get_active_submodule_name(module) != name:
            for _ in range(2):
                self._hover_and_click(module, menu_name, name)
                try:
                    WebDriverWait(submodule, 5).until(lambda ele: ele.get_attribute("style") == "display: none;")
                except TimeoutException:
                    continue
                break
            submodule = self._get_statement_submodule(module)
        time.sleep(5)
        return submodule

    def _get_statement_submodule(self, module: WebElement):
        """获取报表中心激活的子模块"""
        pattern = (By.XPATH, ".//div[@id='__root_wrapper_rms-report']//div[@style='display: block;']")
        WebDriverWait(module, 60).until(EC.presence_of_all_elements_located(pattern))
        current_submodule = module.find_element(*pattern)
        return current_submodule

    def _enter_mc_module(self, module: WebElement, menu_name, name, url):
        """进入营销中心的子模块"""
        print(f"进入营销中心-{menu_name}-{name}")
        if self._get_active_submodule_name(module) != name:
            self._hover_and_click(module, menu_name, name)
            time.sleep(5)
        pattern = (By.XPATH, f".//iframe[@data-current-url='{url}']")
        WebDriverWait(module, 120).until(EC.visibility_of_element_located(pattern))
        iframe = module.find_element(*pattern)
        self._driver.switch_to.frame(iframe)
        time.sleep(2)
        return self._driver

    def _get_active_submodule_name(self, module: WebElement):
        """获取激活的子模块名字"""
        pattern = (By.XPATH, ".//div[@role='tablist']//div[@aria-selected='true']")
        cur_submodule_name = module.find_element(*pattern).text
        return cur_submodule_name

    def _hover_and_click(self, module: WebElement, menu_name, name):
        """悬停菜单栏并点击子模块"""
        pattern = (By.XPATH, f".//div[@class='menu-container ']//span[text()='{menu_name}']/../..")
        WebDriverWait(module, 10).until(EC.visibility_of_element_located(pattern))
        menu = module.find_element(*pattern)
        menu_id = menu.get_attribute("id").split("_")[1]
        for _ in range(4):
            try:
                self._action.move_to_element(menu).perform()
                pattern = (By.XPATH, f"//ul[@id='{menu_id}$Menu']//li[text()='{name}']")
                WebDriverWait(self._driver, 10).until(EC.visibility_of_element_located(pattern))
                self._driver.find_element(*pattern).click()
            except TimeoutException:
                continue
            break

    def _wait_shadow_dom(self) -> WebElement:
        """等待shadow DOM加载完成"""
        WebDriverWait(self._driver, 30).until(EC.visibility_of_element_located((By.TAG_NAME, "render-box-root-x")))  # 等待shadowRoot显现
        time.sleep(3)

    def _js_span_find(self, value):
        """"通过JS命令查找span元素"""
        return f"Array.from({self._js_shadow_root()}.querySelectorAll('span')).find(span => span.textContent.trim() === '{value}')"
    
    def _js_shadow_root(self):
        """查找shadowRoot"""
        return """
            Array.from(document.querySelectorAll('render-box-root-x'))
            .find(root => root.offsetParent !== null).shadowRoot
        """

    def _click_normal_by_js(self, value, text=None, is_parent=False):
        """通用的点击方式,通过CSS选择器定位元素并点击"""
        if text is None:
            ele_pattern = f"const ele = root.shadowRoot.querySelector('{value}')"
        else:
            ele_pattern = f"const ele = Array.from(root.shadowRoot.querySelectorAll('{value}')).find(ele => ele.textContent.trim() === '{text}')"
        if is_parent:
            ele_pattern += ".parentElement"
        return f"""
        const root = Array.from(document.querySelectorAll('render-box-root-x')).find(root => root.offsetParent !== null);
        {ele_pattern};
        if (!ele) return false;
        ele.click();
        return true
        """
    
    def _wait_by_js(self, css_value, text=None, key_value=None, is_has="true", is_parent=False):
        """通用的JS等待显示方式,通过CSS选择器来查找元素
        
        Args:
            css_value: CSS选择器
            key_value: 判断的值,例如hidden
            is_has: 是否是含有,还是不含有
        """
        if text is None:
            ele_pattern = f"const ele = root.shadowRoot.querySelector('{css_value}')"
        else:
            ele_pattern = f"const ele = Array.from(root.shadowRoot.querySelectorAll('{css_value}')).find(ele => ele.textContent.trim() === '{text}')"
        if is_parent:
            ele_pattern += ".parentElement"
        if key_value is None:
            key_pattern = "return true"
        else:
            key_pattern = f"""
            const hasKey = Array.from(ele.classList).some(cls => cls.includes('{key_value}'));
            return hasKey === {is_has};
            """
        return f"""
        const root = Array.from(document.querySelectorAll('render-box-root-x')).find(ele => ele.offsetParent !== null);
        {ele_pattern};
        if (!ele) return false;
        {key_pattern}
        """

    def _wait_search_btn_by_js(self, is_finish=True):
        """判断查询按钮是否点击,以及是否完成点击(即内容是否显示)"""
        return f"""
        const root = Array.from(document.querySelectorAll('render-box-root-x')).find(ele => ele.offsetParent !== null);
        const span = root.shadowRoot.querySelector('button.saas-btn-primary span.saasicon-search');
        const parent = span.parentElement
        const val = parent.getAttribute('saas-click-animating-without-extra-node')
        if (val {'=' if is_finish else '!'}== 'false') return true;
        return false
        """

    def _wait_export(self, value):
        """判断是否处于导出加载状态"""
        return f"""
        const root = Array.from(document.querySelectorAll('render-box-root-x')).find(ele => ele.offsetParent !== null);
        const span = Array.from(root.shadowRoot.querySelectorAll('{value}')).find(ele => ele.textContent.trim() === '导出');
        const button = span.parentElement
        if (button.className.includes('loading')) return true;
        return false
        """

    def _js_click(self, action, signal, timeout=3):
        """通过JS命令点击元素"""
        for _ in range(10):
            result = self._driver.execute_script(action)
            if not result:
                time.sleep(2)
                continue
            if self._js_wait(signal, timeout):
                return True
            else:
                continue
        raise Exception(f"多次尝试点击操作失败:{action}")

    def _js_wait(self, script, timeout=3):
        """通过JS等待元素显示"""
        st = time.time()
        while time.time() - st < timeout:
            result = self._driver.execute_script(script)
            if result:
                return True
            else:
                time.sleep(0.5)
                continue
        return False

    def _js_execute_result(self, value):
        """封装JS命令,执行失败报错"""
        value = f"""
            const resultVariable = {value};
            if (!resultVariable) return false;
            return true
        """
        return value

    def _js_click_start_date(self):
        """"通过JS点击shadowRoot里面的开始日期控件"""
        action = f"""
        {self._js_shadow_root()}.querySelector('input[placeholder=\"开始日期\"]').parentElement.click();
        return true
        """
        signal = f"!{self._js_span_find('上月')}.closest('div.saas-picker-dropdown-range').className.includes('saas-picker-dropdown-hidden')"
        self._js_click(action, self._js_execute_result(signal))

    def _js_click_last_month(self):
        """"通过JS点击shadowRoot里日期控件中的上月按钮"""
        action = f"{self._js_span_find('上月')}.click() \n return true"
        signal = f"{self._js_span_find('上月')}.closest('div.saas-picker-dropdown-range').className.includes('saas-picker-dropdown-hidden')"
        self._js_click(action, self._js_execute_result(signal))
    
    def _js_click_search(self):
        """"通过JS点击shadowRoot里的查询按钮"""
        action = f"{self._js_span_find('查询')}.click() \n return true"
        signal = f"{self._js_shadow_root()}.querySelector('li[title=\"上一页\"]')"
        self._js_click(action, self._js_execute_result(signal), timeout=5)
    
    def _js_click_download(self):
        """"通过JS点击shadowRoot里的下载按钮"""
        action = f"{self._js_span_find('导出')}.click() \n return true"
        signal = f"{self._js_span_find('导出')}.parentNode.className.includes('saas-btn-loading') \n return true"
        self._js_click(action, self._js_execute_result(signal))

    def _date_select_1(self, submodule: WebElement):
        """日期选择1:支付结算,支付明细"""
        pattern = (By.XPATH, ".//input[@placeholder='请选择日期']")
        WebDriverWait(submodule, 60).until(EC.element_to_be_clickable(pattern))
        submodule.find_element(*pattern).click()
        pattern = (By.XPATH, "//div[@class='ant-calendar-footer']//span[text()='上月']")
        WebDriverWait(self._driver, 10).until(EC.visibility_of_element_located(pattern))
        self._driver.find_element(*pattern).click()
        WebDriverWait(self._driver, 10).until_not(EC.visibility_of_element_located(pattern))

    def _date_select_3(self, submodule: WebElement):
        """日期选择3:会员新增情况统计表的那个类型日期选择控件"""
        condition = EC.element_to_be_clickable((By.XPATH, ".//input[@placeholder='请选择日期']"))
        select = WebDriverWait(submodule, 10).until(condition)
        select.click()
        WebDriverWait(self._driver, 10).until(EC.visibility_of_element_located((By.CLASS_NAME, "ant-calendar")))
        last_month, calendar_ele = self._locate_last_month()
        self._driver.find_element(By.XPATH, f".//td[@title='{last_month}1日']").click()
        WebDriverWait(self._driver, 10).until(EC.invisibility_of_element(calendar_ele))
        last_month, calendar_ele = self._locate_last_month()
        pattern = (By.XPATH, ".//td[@class='ant-calendar-cell ant-calendar-last-day-of-month']")
        self._driver.find_element(*pattern).click()
        WebDriverWait(self._driver, 10).until(EC.invisibility_of_element(calendar_ele))

    def _date_select_4(self, submodule: WebElement):
        """日期选择4:储值消费汇总表（新版）的那个类型日期选择控件"""
        WebDriverWait(submodule, 60).until(EC.visibility_of_element_located((By.XPATH, "//div[text()='自然日']")))
        start_ele, end_ele = submodule.find_elements(By.CLASS_NAME, "saas-picker-input")
        now_date = datetime.today()
        end_date = now_date.replace(day=1) - relativedelta(days=1)
        start_date = end_date.replace(day=1)
        for ele, select_date in zip([start_ele, end_ele], [start_date, end_date]):
            ele.click()
            label = WebDriverWait(submodule, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, "saas-picker-header-view")))
            label_date = datetime.strptime(label.text, "%Y年%m月")
            last_year_btn = submodule.find_element(By.CLASS_NAME, "saas-picker-header-super-prev-btn")
            next_year_btn = submodule.find_element(By.CLASS_NAME, "saas-picker-header-super-next-btn")
            last_month_btn = submodule.find_element(By.CLASS_NAME, "saas-picker-header-prev-btn")
            next_month_btn = submodule.find_element(By.CLASS_NAME, "saas-picker-header-next-btn")
            while True:
                label = submodule.find_elements(By.CLASS_NAME, "saas-picker-header-view")[0]
                label_date = datetime.strptime(label.text, "%Y年%m月")
                if label_date.year == select_date.year and label_date.month == select_date.month:
                    break
                if label_date.year > select_date.year:
                    last_year_btn.click()
                elif label_date.year < select_date.year:
                    next_year_btn.click()
                elif label_date.month > select_date.month:
                    last_month_btn.click()
                elif label_date.month < select_date.month:
                    next_month_btn.click()
                WebDriverWait(label, 10).until(lambda ele: ele.text != label_date.strftime("%Y年%m月") and ele.text != "")
            select_date_str = select_date.strftime('%Y-%m-%d')
            submodule.find_element(By.XPATH, f"//td[@title='{select_date_str}']").click()
        submodule.find_element(By.XPATH, "//span[text()='确 定']").click()

    def _locate_last_month(self):
        """日期选择3的日历控件定位到上个月"""
        now_date = datetime.today()
        while True:
            calendar_month_ele = self._driver.find_element(By.CLASS_NAME, "ant-calendar-ym-select")
            calendar_month_str = calendar_month_ele.text
            year_int, month_int = calendar_month_str.split("年")
            year_int = int(year_int)
            month_int = int(month_int[:-1])
            if year_int == now_date.year and month_int == now_date.month:
                self._driver.find_element(By.CLASS_NAME, "ant-calendar-prev-month-btn").click()
                WebDriverWait(calendar_month_ele, 10).until(lambda ele: ele.text != calendar_month_str)
                break
            calendar_month = datetime.strptime(calendar_month_str, '%Y年%m月')
            if now_date - calendar_month <= timedelta(0):
                self._driver.find_element(By.CLASS_NAME, "ant-calendar-prev-month-btn").click()
            else:
                self._driver.find_element(By.CLASS_NAME, "ant-calendar-next-month-btn").click()
            WebDriverWait(calendar_month_ele, 60).until(lambda ele: ele.text != calendar_month_str)
        return calendar_month_ele.text, calendar_month_ele

    def _search(self, submodule: WebElement):
        ele = WebDriverWait(submodule, 60).until(EC.element_to_be_clickable((By.XPATH, ".//button[contains(., '查询')]")))
        ele.click()
        time.sleep(5)

    def _wait_search_finnsh_1(self, submodule: WebElement):
        """等待查询内容显示1:类似综合收款统计的表格类型"""
        pattern = (By.XPATH, ".//div[@class='ant-spin-nested-loading']//div[@class='ant-spin-container']")
        WebDriverWait(submodule, 60).until(EC.presence_of_element_located(pattern))

    def _wait_search_finnsh_2(self):
        """等待查询内容显示2:类似储值消费汇总表的表格类型"""
        condition = EC.presence_of_element_located((By.CLASS_NAME, 'el-loading-parent--relative'))
        WebDriverWait(self._driver, 60).until_not(condition)

    def _clear_excel(self, name):
        """清理已存在的文件"""
        file_names = os.listdir(self._download_path)
        for file_name in file_names:
            if name not in file_name:
                continue
            path = os.path.join(self._download_path, file_name)
            os.remove(path)
            print(f"清理文件:{path}")

    def _download_direct(self, submodule: WebElement, download_name, save_name):
        """直接导出文件"""
        submodule.find_element(By.XPATH, ".//span[text()='导出']/parent::button").click()
        self.wait_download(f"*{download_name}*", save_name)

    def _download_autotrophy_detail(self, submodule: WebElement, name):
        """导出文件"""
        submodule.find_element(By.XPATH, ".//span[text()='导出']/parent::button").click()
        condition = EC.visibility_of_element_located((By.XPATH, "//div[@id='rcDialogTitle0']/../.."))
        dialog = WebDriverWait(self._driver, 10).until(condition)
        for content in ["菜品明细", "支付明细", "优惠明细"]:
            select_ele = dialog.find_element(By.XPATH, f".//span[text()='{content}']/preceding-sibling::span")
            if "ant-checkbox-checked" not in select_ele.get_attribute("class"):
                select_ele.click()
            WebDriverWait(select_ele, 2).until(lambda ele: "ant-checkbox-checked" in ele.get_attribute("class"))
        dialog.find_element(By.XPATH, ".//span[text()='确 定']/parent::button").click()
        self.wait_download(f"*{name}*", name)

    def _pay_settlement_condition(self, submodule: WebElement):
        """支付结算表的查询条件"""
        element = submodule.find_element(By.XPATH, ".//span[text()='交易日期']/..")
        if "isSelected" in element.get_attribute("class"):
            print("当前统计方式已经是：交易日期")
        else:
            element.click()

    def _toggle_old(self):
        """切换到旧版本:会员新增情况统计表的功能"""
        # 收缩导航
        self._driver.switch_to.default_content()
        ele = WebDriverWait(self._driver, 30).until(EC.element_to_be_clickable((By.ID, "cs-entry-icon")))
        if "cs-entry-logo_unfolded" in ele.get_attribute("class"):
            ele.click()
            WebDriverWait(self._driver, 30).until(lambda d: "cs-entry-logo_folded" in ele.get_attribute("class"))
        pattern = (By.XPATH, f".//iframe[@data-current-url='/web/member/statistic/member-increase#/']")
        WebDriverWait(self._driver, 60).until(EC.visibility_of_element_located(pattern))
        iframe = self._driver.find_element(*pattern)
        self._driver.switch_to.frame(iframe)
        c1 = EC.visibility_of_element_located((By.XPATH, "//span[text()='切换回老版']"))
        c2 = EC.visibility_of_element_located((By.XPATH, "//span[text()='切换回新版']"))
        version = WebDriverWait(self._driver, 60).until(EC.any_of(c1, c2))
        if version.text == "切换回老版":
            version.click()
        else:
            print("当前已经在老版本")


class ElemeData:
    """饿了么相关留存表的处理"""

    def __init__(self) -> None:
        self.wb = load_workbook(GOL.save_path.eleme_bill)

    def save(self):
        self.wb.save(GOL.save_path.eleme_bill)

    def del_useless_sheet(self):
        """删除没用的分表"""
        for name in self.wb.sheetnames:
            if name in ["账单汇总", "外卖账单明细", "抖音渠道佣金明细","保险相关业务账单明细", "赔偿单", "客单价"]:
                continue
            print(f"删除分表:{name}")
            sheet = self.wb[name]
            self.wb.remove(sheet)

    def adjust_font_size(self):
        """调整表格字体大小"""
        font_11 = Font(size=11)
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font_11

    def billing_summary(self):
        """处理分表-账单汇总"""
        ws = self.get_ws("账单汇总")
        header = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        assert header == ['结算入账ID', '门店ID', '门店名称', '账单日期', '结算金额', '结算日期', '账单类型']
        # 将数据转化为pandas
        df = pd.DataFrame(list(ws.values)[1:], columns=header)
        df_result = df[df["账单类型"] == "外卖"]
        # 插入数据: 保障服务费
        insurance_data = df[df["账单类型"] == "保障服务费"]
        insurance_data = insurance_data[["账单日期", "结算金额"]]
        insurance_data = insurance_data.rename(columns={"结算金额": "保障服务费"})
        df_result = insert_data_column_merge(df_result, insurance_data, "结算金额", "保障服务费")
        # 插入数据: 闪购联盟推广
        eleme_data = df[df["账单类型"] == "闪购联盟推广"]
        if len(eleme_data) == 0:
            print("数据为空,请注意!!!------闪购联盟推广")
        else:
            eleme_data = eleme_data[["账单日期", "结算金额"]]
            eleme_data = eleme_data.rename(columns={"结算金额": "闪购联盟推广"})
            df_result = insert_data_column_merge(df_result, eleme_data, "结算金额", "闪购联盟推广")
        # 插入数据：全站推广
        promotion_data = df[df["账单类型"] == "全站推广"]
        if len(promotion_data) == 0:
            print("数据为空,请注意!!!------全站推广")
        else:
            promotion_data = promotion_data[["账单日期", "结算金额"]]
            promotion_data = promotion_data.rename(columns={"结算金额": "全站推广"})
            df_result = insert_data_column_merge(df_result, promotion_data, "结算金额", "全站推广")
        # 插入数据: 抖音渠道佣金
        tiktok_ws = self.get_ws("抖音渠道佣金明细") 
        if tiktok_ws is not None and tiktok_ws.max_row > 1:
            header = [tiktok_ws.cell(1, i).value for i in range(1, tiktok_ws.max_column + 1)]
            date_i = header.index("账单日期")
            amount_i = header.index("结算金额")
            tiktok_data = {}
            for row in tiktok_ws.iter_rows(min_row=2, values_only=True):
                date_str = row[date_i]
                amount = row[amount_i]
                tiktok_data[date_str] = amount + tiktok_data.get(date_str, 0)
            tiktok_data = pd.DataFrame(list(tiktok_data.items()), columns=["账单日期", "抖音渠道佣金"])
            df_result = insert_data_column_merge(df_result, tiktok_data, "结算金额", "抖音渠道佣金")
        # 插入数据：小计
        existing_columns = [col for col in ["结算金额", "抖音渠道佣金", "闪购联盟推广", "保障服务费"] if col in df_result.columns]
        total = df_result[existing_columns].fillna(0).sum(axis=1)
        insert_i = list(df_result.columns).index("结算日期")
        df_result.insert(insert_i, "小计", total)
        # 写入标题
        df_result_header = list(df_result.columns)
        header_style = ws["A1"]._style
        for j, value in enumerate(df_result_header):
            each_cell = ws.cell(1, j + 1)
            each_cell.value = value
            each_cell._style = copy(header_style)
        # 写入数据
        data_row = len(df_result)
        for i, row in df_result.iterrows():
            for j, value in enumerate(row):
                ws.cell(i + 2, j + 1, value)
        ws.delete_rows(data_row+2, ws.max_row - data_row-1) 
        # 对数据进行求和
        row_i = len(df_result) + 2
        header = df_result.columns.tolist()
        for i in range(header.index("结算金额"), header.index("小计")+1):
            ws.cell(row_i, i+1, df_result[header[i]].sum())
        # 给所有数据加框线
        thin_side  = Side(border_style="thin", color="000000")
        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

    def get_ws(self, name):
        """获取工作表,排除空格等障碍"""
        for sheet_name in self.wb.sheetnames:
            if sheet_name.replace(" ", "") == name:
                return self.wb[sheet_name]

    def take_out(self):
        """处理分表-外卖账单明细"""
        ws = self.wb["外卖账单明细"]
        header = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        # 插入标题
        col_i = header.index("结算金额") + 2
        ws.insert_cols(col_i, 2)
        ws.cell(1, col_i, "结算")
        ws.cell(1, col_i+1, "差额")
        header = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        # 修改格式
        number_style = NamedStyle(name="number_style", number_format="0.00")
        for col in range(column_index_from_string("J"), ws.max_column + 1):
            for row in range(1, ws.max_row+1):
                ws.cell(row, col).style = number_style
        ws.freeze_panes = "A2"
        # 填充列数据---结算
        settle_letters = ['Q', 'R', 'T', 'AE', 'AF', 'AG', 'AJ', 'BE', 'BF', 'BH', 'BW', 'BX', 'BY', 'BZ', 'CC', 'AT']
        settle_col_i = header.index("结算") + 1
        for i in range(2, ws.max_row + 1):
            value = [f"{letter}{i}" for letter in settle_letters]
            ws.cell(i, settle_col_i, f"={'+'.join(value)}")
        # 填充"差额"列的公式
        diff_col_i = header.index("差额") + 1 
        for i in range(2, ws.max_row + 1):
            ws.cell(i, diff_col_i, "=K3-J3")
        # 创建新表---赔偿单：差额不为0的数据
        settle_indexs = [column_index_from_string(letter) for letter in settle_letters]
        compensate_ws: Worksheet = None
        row_i = 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            settle_value = [Decimal(str(row[i-1])) for i in settle_indexs]  # 计算数据
            balance = sum(settle_value) - Decimal(str(row[settle_col_i-2]))
            if balance == Decimal(str(0)):
                continue
            if compensate_ws is None:
                compensate_ws = self.wb.create_sheet("赔偿单")
                for i in range(1, ws.max_column + 1):
                    compensate_ws.cell(1, i, ws.cell(1, i).value)
            for i, value in enumerate(row):
                compensate_ws.cell(row_i, i+1, value)
            row_i += 1
        if compensate_ws is not None:
            self.add_total_row(compensate_ws)
        # 创建新表---客单价：订单子类型为即时单
        value_i = header.index("订单子类型")
        custom_ws: Worksheet = None
        row_i = 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[value_i] != "即时单":
                continue
            if custom_ws is None:
                custom_ws = self.wb.create_sheet("客单价")
                for i in range(1, ws.max_column + 1):
                    value = ws.cell(1, i).value
                    custom_ws.cell(1, i, value)
            for i, value in enumerate(row):
                custom_ws.cell(row_i, i+1, value)
            row_i += 1
        if custom_ws is not None:
            self.add_total_row(custom_ws)
        # 最后一行新增合计数
        self.add_total_row(ws) 
        sum_row_i = ws.max_row + 1
        ws.cell(sum_row_i, 1, "合计")
        for i in range(column_index_from_string("J"), column_index_from_string("CC")):
            col_str = get_column_letter(i)
            range_start = f"{col_str}2"
            range_end = f"{col_str}{sum_row_i - 1}"
            ws.cell(sum_row_i, i, f"=SUM({range_start}:{range_end})")    
            ws.column_dimensions[col_str].width = 12
    
    def add_total_row(self, ws: Worksheet):
        """新增合计行"""
        sum_row_i = ws.max_row + 1
        ws.cell(sum_row_i, 1, "合计")
        for i in range(10, ws.max_column + 1):
            col_str = get_column_letter(i)
            range_start = f"{col_str}2"
            range_end = f"{col_str}{sum_row_i - 1}"
            ws.cell(sum_row_i, i, f"=SUM({range_start}:{range_end})")    

class DaDaAutotrophy:
    """达达门店订单明细的处理"""

    def __init__(self) -> None:
        self.wb = load_workbook(GOL.save_path.autotrophy_dada)

    def save(self):
        self.wb.save(GOL.save_path.autotrophy_dada)

    def extract_data(self):
        """从原表中提取所需数据"""
        # 获取需要保存的数据
        ws = self.wb["1"]
        header = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        order_source_i = header.index("订单来源编号")
        order_status_i = header.index("订单状态")
        autotrophy = [header]
        anomaly = [header]
        for row in ws.iter_rows(min_row=2, values_only=True):
            order_source = row[order_source_i]
            order_status = row[order_status_i]
            if order_source is None or "自营外卖" not in order_source:
                continue
            if order_status == '已完成':
                autotrophy.append(list(row))
            elif order_status in ["已取消", "妥投异常,返还完成"]:
                anomaly.append(list(row))
        return autotrophy, anomaly

    def proper_anomaly(self, datas):
        """新建妥投异常分表"""
        ws: Worksheet = self.wb.create_sheet("妥投异常及取消订单运费")
        for i, row in enumerate(datas, start=1):
            for j, value in enumerate(row, start=1):
                ws.cell(i, j, value)
        # 添加合计行
        sum_row_i = ws.max_row + 1
        ws.cell(sum_row_i, 1, "合计")
        for i in range(15, 41):
            range_start = f"{get_column_letter(i)}2"
            range_end = f"{get_column_letter(i)}{sum_row_i - 1}"
            ws.cell(sum_row_i, i, f"=SUM({range_start}:{range_end})")

    def autotrophy_order(self, datas: List[list]):
        """新建自营订单外卖分表"""
        header = datas[0]
        # 将配送距离的值从文本改为数值
        distance_i = header.index("配送距离")
        for i in range(1, len(datas)):
            datas[i][distance_i] = float(datas[i][distance_i])
        # 保存数据
        ws: Worksheet = self.wb.create_sheet("自营外卖订单（不含自配送）")
        for i, row in enumerate(datas, start=1):
            for j, value in enumerate(row, start=1):
                ws.cell(i, j, value)
        # 插入数据
        ws.insert_cols(15, 1)
        ws.cell(1, 15, "配送区间应收客户运费")
        assert ws.cell(1, 14).value == "配送距离"
        for i in range(2, ws.max_row + 1):
            formula = f"=IF(N{i}<=2000,2,IF(N{i}<=3000,3,IF(N{i}<=4000,4,IF(N{i}<=5000,6,IF(N{i}<=6000,7,"
            formula += f"IF(N{i}<=7000,9,IF(N{i}<=9000,10,IF(N{i}<=10000,12,IF(N{i}<=12000,15,"
            formula += f"IF(N{i}<=14000,18,IF(N{i}<=15000,22,"")))))))))))"
            ws.cell(i, 15, formula)
        ws.insert_cols(40, 1)
        ws.cell(1, 40, "商户支付配送费")
        assert ws.cell(1, 39).value == "运费账户消耗"
        for i in range(2, ws.max_row + 1):
            ws.cell(i, 40, f"=AM{i}-O{i}")
        # 添加合计行
        sum_row_i = ws.max_row + 1
        ws.cell(sum_row_i, 1, "合计")
        for i in range(15, 41):
            range_start = f"{get_column_letter(i)}2"
            range_end = f"{get_column_letter(i)}{sum_row_i - 1}"
            ws.cell(sum_row_i, i, f"=SUM({range_start}:{range_end})")


class MeiTuanAutotrophy:
    """美团自营外卖表的处理"""

    def __init__(self) -> None:
        self.wb = load_workbook(GOL.save_path.autotrophy_meituan)

    def save(self):
        self.wb.save(GOL.save_path.autotrophy_meituan)

    def order_detail(self):
        """订单明细分表的处理"""
        ws = self.wb["订单明细"]
        # 整理格式
        merge_range_cells = list(ws.merged_cells.ranges)
        for merge_range in merge_range_cells:
            ws.unmerge_cells(str(merge_range))
        ws.delete_rows(1, 2)
        ws.freeze_panes = "A2"
        # 处理订单明细分表
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 13
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for col in list("CEFQ"):
            cell: Cell = ws[f"{col}1"]
            cell.fill = yellow_fill
        # 获取需要另存为的数据
        header = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        order_source_i = header.index("订单来源")
        order_status_i = header.index("订单状态")
        distribution_mode_i = header.index("配送方式")
        save_data = [header]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[order_source_i] != "自营外卖":
                continue
            elif row[order_status_i] != "已完成":
                continue
            elif row[distribution_mode_i] != "自配送(达达配送)":
                continue
            save_data.append(row)
        # 保存数据
        ws = self.wb.create_sheet("自营外卖订单明细（不含自配送）")
        for i, row in enumerate(save_data, start=1):
            for j, value in enumerate(row, start=1):
                ws.cell(i, j, value)
        # 添加合计行
        sum_row_i = ws.max_row + 1
        ws.cell(sum_row_i, 1, "合计")
        sum_term_index = [header.index(name) for name in ['订单金额（元）', '顾客应付（元）', '支付合计（元）', '订单优惠（元）', '订单收入（元）']]
        for i in sum_term_index:
            i = i + 1
            col_str = get_column_letter(i)
            range_start = f"{col_str}2"
            range_end = f"{col_str}{sum_row_i - 1}"
            ws.cell(sum_row_i, i, f"=SUM({range_start}:{range_end})")


class TalkOutData:
    """外卖订单汇总表的数据处理"""

    def __init__(self) -> None:
        self.wb = load_workbook(GOL.save_path.take_out)

    def save(self):
        self.wb.save(GOL.save_path.take_out)

    def collect_eleme(self):
        """汇总饿了么的数据"""
        # 为了获取公式计算后的结果
        with xw.App(visible=False) as app:
            with app.books.open(GOL.save_path.eleme_bill) as app_wb:
                app_ws: Sheet = app_wb.sheets["外卖账单明细"]
                last_cell = app_ws.used_range.last_cell
                max_row, max_col = last_cell.row, last_cell.column
                header: List = app_ws.range((1, 1, 1, max_col)).value
                should_income = app_ws.range(max_row, header.index("商品金额")).value # 应收
                service_cost = app_ws.range(max_row, header.index("技术服务费")).value  # 抽佣服务费
                time_raise = app_ws.range(max_row, header.index("时段收费")).value * 0.85  # 时段加价
                distance_raise = app_ws.range(max_row, header.index("距离收费")).value * 0.85  # 距离加价
                price_raise = (app_ws.range(max_row, header.index("配送服务费")).value - time_raise - distance_raise) * 0.85  # 价格加价
                activty_subsidy = sum([app_ws.range(max_row, header.index(name)).value 
                                        for name in ["商家活动补贴", "商家代金券补贴", "智能满减津贴"]])  # 商户承担活动补贴
                chargeback = - app_ws.range(max_row, header.index("差额")).value + app_ws.range(
                    max_row, header.index("打包费")).value  # 部分退单-申请退单金额
                other = sum([app_ws.range(max_row, header.index(name)).value for 
                                name in ["商家配送费活动补贴", "商家呼单小费"] if name in header])  # 其他(商家自行配送补贴
                eat_now_pay_later = app_ws.range(max_row, header.index("先享后付服务费")).value  # 先享后付服务费
                app_ws = app_wb.sheets["账单汇总"]
                last_cell = app_ws.used_range.last_cell
                max_row, max_col = last_cell.row, last_cell.column
                header: List = app_ws.range((1, 1, 1, max_col)).value
                insurance = 0  # 保险费
                if "保险费" in header:
                    insurance = insurance + app_ws.range(max_row, header.index("保险金额")).value
                if "抖音渠道佣金" in header:
                    insurance = insurance - app_ws.range(max_row, header.index("抖音渠道佣金")).value
        # 写入饿了么数据表
        ws = self.wb[f"{GOL.last_month.year}年饿了么"]
        row_i = [str(ws.cell(i, 1).value) for i in range(4, 16)]
        row_i = row_i.index(GOL.last_month.strftime("%y.%m")) + 4
        ws.cell(row_i, 2, should_income)  
        ws.cell(row_i, 3, service_cost)  
        ws.cell(row_i, 4, time_raise)  
        ws.cell(row_i, 5, distance_raise)  
        ws.cell(row_i, 6, price_raise)  
        ws.cell(row_i, 7, activty_subsidy)  
        ws.cell(row_i, 8, chargeback)  
        ws.cell(row_i, 9, other)  
        ws.cell(row_i, 10, eat_now_pay_later)
        ws.cell(row_i, 12, insurance) 

    def collect_autotrophy(self):
        """汇总自营外卖的数据"""
        with xw.App(visible=False) as app:
            with app.books.open(GOL.save_path.autotrophy_meituan) as wb:
                ws = wb.sheets["自营外卖订单明细（不含自配送）"]
                last_cell = ws.used_range.last_cell
                max_row, max_col = last_cell.row, last_cell.column
                header: List = ws[f"A1:{get_column_letter(max_col)}1"].value

                order_num = max_row - 2
                col_str = get_column_letter(header.index("结账方式") + 1)

                member_consume = len([data for data in ws[f"{col_str}1:{col_str}{max_row - 1}"].value if "会员卡" in data])
                order_amount = ws[f"{get_column_letter(header.index('订单金额（元）') + 1)}{max_row}"].value
                discount_amount = ws[f"{get_column_letter(header.index('订单优惠（元）') + 1)}{max_row}"].value
            with app.books.open(GOL.save_path.autotrophy_dada) as wb:
                ws = wb.sheets["自营外卖订单（不含自配送）"]
                last_cell = ws.used_range.last_cell
                max_row, max_col = last_cell.row, last_cell.column
                header: List = ws[f"A1:{get_column_letter(max_col)}1"].value

                client_freight = ws[f"{get_column_letter(header.index('配送区间应收客户运费') + 1)}{max_row}"].value
                business_freight = ws[f"{get_column_letter(header.index('商户支付配送费') + 1)}{max_row}"].value

                ws = wb.sheets["妥投异常及取消订单运费"]
                last_cell = ws.used_range.last_cell
                max_row, max_col = last_cell.row, last_cell.column
                header: List = ws[f"A1:{get_column_letter(max_col)}1"].value

                business_freight += ws[f"{get_column_letter(header.index('运费账户消耗') + 1)}{max_row}"].value

        ws = self.wb[f"{GOL.last_month.year}年自营外卖"]
        row_i = [str(ws.cell(i, 1).value) for i in range(3, 15)]
        row_i = row_i.index(GOL.last_month.strftime("%Y.%m")[2:]) + 3
        ws.cell(row_i, 2, order_num)  # 订单数量
        ws.cell(row_i, 3, member_consume)  # 会员消费订单数
        ws.cell(row_i, 5, order_amount)  # 订单金额
        ws.cell(row_i, 6, discount_amount)  # 折扣金额(含会员、运费及商品）
        ws.cell(row_i, 7, client_freight)  # 客人支付运费
        ws.cell(row_i, 8, business_freight)  # 商户承担运费

def insert_data_column_merge(df_main: pd.DataFrame, df_insurance, insert_after, insert_name):
    """
    使用merge方法插入某列数据
    
    Args:
        df_main: 主DataFrame
        df_insurance: 包含账单日期和所需数据的DataFrame
        insert_after: 在哪个列后面插入
        insert_name: 插入数据的列名
    
    Returns:
        插入数据后的DataFrame
    """
    # 左连接合并，保留所有主表数据
    df_main = df_main.merge(
        df_insurance,
        on='账单日期',
        how='left'
    )
    
    # 获取插入位置
    insert_idx = df_main.columns.get_loc(insert_after) + 1
    
    # 获取'保险金额'列
    insurance_col = df_main.pop(insert_name)
    
    # 插入到指定位置
    df_main.insert(insert_idx, insert_name, insurance_col)
    
    return df_main


def get_3row_index(row1: pd.Series, row2: pd.Series, row3: pd.Series, name1, name2, name3, is_must=True):
    """获取3行标题中某个数据的索引"""
    row1_drop = row1.dropna()
    row1_values = list(row1_drop.values)
    row1_indexs = [*list(row1_drop.index), len(row1)]
    n1_i = row1_values.index(name1)
    n1_s, n1_e = row1_indexs[n1_i], row1_indexs[n1_i + 1]
    row2_drop = row2.iloc[n1_s:n1_e].dropna()
    row2_values = list(row2_drop.values)
    row2_indexs = [*list(row2_drop.index), len(row2)]
    if name2 not in row2_values:
        if is_must:
            raise ValueError(f"{name2} not in {row2_values}")
        return None
    n2_i = row2_values.index(name2)
    n2_s, n2_e = row2_indexs[n2_i], row2_indexs[n2_i + 1]
    if name3 is None:
        return n2_s, n2_e
    if name2 == name3:
        return n2_s
    row3_drop = row3.iloc[n2_s:n2_e]
    row3_values = list(row3_drop.values)
    row3_indexs = list(row3_drop.index)
    return get_value_index(row3_indexs, row3_values, name3, is_must)


def get_2row_index(row1: pd.Series, row2: pd.Series, name1, name2, is_must=True):
    """"获取两行标题的某个数据的索引"

    Returns:
        int: 指定的数据的列索引
    """
    name1_range = get_row_range(row1, name1)
    if name1_range is None and not is_must:
        return None
    n1_s, n1_e = name1_range
    row2_drop = row2.iloc[n1_s:n1_e].dropna()
    row2_values = list(row2_drop.values)
    row2_indexs = list(row2_drop.index)
    return get_value_index(row2_indexs, row2_values, name2, is_must)

def get_row_range(row:pd.Series, name):
    """获取某个数据在该行的索引
    
    Returns:
        int: 开始的索引值 
        int: 结束的索引值
    """
    row_drop = row.dropna()
    row_values = list(row_drop.values)
    row_indexs = [*list(row_drop.index), len(row)]
    if name not in row_values:
        return None
    n1_i = row_values.index(name)
    n1_s, n1_e = row_indexs[n1_i], row_indexs[n1_i + 1]
    return n1_s, n1_e

def get_value_index(indexs, names:list, search_names, is_must):
    """获取多个标题中某个数据的索引"""
    if isinstance(search_names, str):
        if search_names in names:
            return indexs[names.index(search_names)]
        if is_must:
            raise ValueError(f"{names}中没有找到{search_names}的数据")
        return None
    else:
        name3_intersect = list(set(names).intersection(set(search_names)))
        if len(name3_intersect) != 0:
            return indexs[names.index(name3_intersect[0])]
        if is_must:
            raise ValueError(f"{names}中没有找到{search_names}的数据")
        return None

def replace_parentheses(text):
    """将半角括号替换成全角括号"""
    if isinstance(text, str):
        text = text.replace('(', '（').replace(')', '）')
    return text

def init_chrome(path, driver_path, download_path, user_path):
    service = Service(driver_path)
    options = Options()
    if path is not None:
        options.binary_location = path
    options.add_argument('--log-level=3')
    options.add_argument(f'user-data-dir={user_path}')
    options.add_experimental_option('prefs', {
        "download.default_directory": download_path,  # 指定下载目录
    })
    chrome_driver = Chrome(service=service, options=options)
    return chrome_driver


def list_generate(indexs, values):
    """list的列表生成器"""
    for i in indexs:
        yield values[i]

def copy_folder(src_folder, dst_folder):
    print("备份网站数据到备份文件夹")
    if os.path.exists(dst_folder):
        shutil.rmtree(dst_folder)
    os.makedirs(dst_folder)
    shutil.copytree(src_folder, dst_folder, dirs_exist_ok=True)


def crawler_main(chrome_path, driver_path, download_path, user_path):
    """网站抓取的主方法"""
    print("启动浏览器，开始爬虫抓取")
    driver = init_chrome(chrome_path, driver_path, download_path, user_path)
    all_download = False
    meituan = MeiTuanCrawler(driver, download_path)
    for path in meituan._name2save.values():
        if os.path.exists(path):
            continue
        break
    else:
        all_download = True
    if not all_download:
        print("登录并打开美团网站")
        meituan.login()
        print("切换店铺")
        meituan.toggle_store(GOL.store_name)
        time.sleep(3)
        print("下载综合营业统计表")
        meituan.download_synthesize_operate()
        print("下载自营外卖/自提订单明细表")
        meituan.download_autotrophy()
        print("下载综合收款统计表的数据")
        meituan.download_synthesize_income()
        print("下载支付结算表的相关数据")
        meituan.download_pay_settlement()
        print("下载支付明细表的相关数据")
        meituan.download_pay_detail()
        print("下载储值消费汇总表的数据")
        meituan.download_store_consume()
        print("下载会员新增情况统计表的相关数据")
        meituan.download_member_addition()
        print("从美团网站爬虫导出EXCEL文件已完成")
    all_download = False
    dada = DadaCrawler(driver, download_path)
    for path in dada._name2save.values():
        if os.path.exists(path):
            continue
        break
    else:
        all_download = True
    if not all_download:
        print("登入并打开达达网站")
        dada.login()
        print("下载门店明细报表")
        dada.download_store_report()
        print("从达达网站爬虫导出EXCEL文件已完成")
    driver.quit()

def operation_detail_main(template_path):
    """营业明细表汇总的主方法"""
    print("定义数据保存类")
    writer = GetOperateDetail(template_path)
    print("读取综合营业统计表的数据")
    writer.read_general_business()
    print("读取综合收款统计表的数据")
    writer.read_general_collection()
    print("读取储值消费汇总表的数据")
    writer.read_store_consume()
    print("读取会员新增情况统计表的相关数据")
    writer.read_newly_increased()
    print("读取支付结算表的相关数据")
    writer.read_pay_settlement()
    print("将数据写入定义好的excel模板中")
    writer.write_and_save()
    print("营业明细表的数据已汇总完成")


def eleme_main():
    """饿了么数据的表格《账单明细》处理的主方法"""
    eleme_data = ElemeData()
    print("处理分表-账单汇总")
    eleme_data.billing_summary()
    print("处理分表-外卖账单明细")
    eleme_data.take_out()
    print("删除不保留的分表")
    eleme_data.del_useless_sheet()
    print("调整整个表格的字体大小")
    eleme_data.adjust_font_size()
    print("保存文件")
    eleme_data.save()
    print("饿了么数据的留存表已整理完毕")


def dada_autotrophy_main():
    """达达门店订单明细处理的主方法"""
    excel = DaDaAutotrophy()
    print("从达达原始数据中提取所需数据")
    data1, data2 = excel.extract_data()
    print("新建自营外卖订单分表")
    excel.autotrophy_order(data1)
    print("新建妥投异常分表")
    excel.proper_anomaly(data2)
    print("保存文件")
    excel.save()
    print("达达门店明细的留存表已整理完毕")


def meituan_autotrophy_main():
    """美团自营外卖表处理的主方法"""
    excel = MeiTuanAutotrophy()
    print("处理订单明细分表")
    excel.order_detail()
    print("保存文件")
    excel.save()
    print("美团自营外卖表的留存表已整理完毕")


def take_out_main():
    """外卖收入汇总表的主方法"""
    take_out = TalkOutData()
    print("汇总饿了么的数据")
    take_out.collect_eleme()
    print("汇总自营外卖的数据")
    take_out.collect_autotrophy()
    print("保存文件")
    take_out.save()
    print("外卖收入的数据已汇总完成")


def main():
    """主方法"""
    operate_detail_template = r"E:\NewFolder\xiayun\营业明细表模板.xlsx"
    user_path = r'C:\Users\Administrator\AppData\Local\Google\Chrome\User Data'
    chrome_driver_path = r'E:\NewFolder\chromedriver_mac_arm64_114\chromedriver.exe'
    chrome_path = r"E:\NewFolder\chromedriver_mac_arm64_114\chrome114\App\Chrome-bin\chrome.exe"
    download_path = r"D:\Download"
    # 定义保存名称
    save_folder = os.path.join(os.path.dirname(operate_detail_template), "输入数据")
    if not os.path.exists(save_folder):
        os.makedirs(save_folder)
    backup_folder = os.path.join(os.path.dirname(operate_detail_template), "数据备份")
    date_str = GOL.last_month.strftime("%Y%m")
    # 输入文件地址, Bread（湖明店）,Bread（瑞景店）,Sweet
    input_value_name = "&&".join(os.listdir(save_folder))
    for deal_name in TM.__dict__.values():
        if deal_name not in input_value_name:
            continue
        print(f"开始处理{deal_name}")
        GOL.store_name = deal_name
        GOL.save_path.take_out = os.path.join(save_folder, f"{GOL.store_name}外卖收入汇总表{date_str}.xlsx")
        GOL.save_path.eleme_bill = os.path.join(save_folder, f"{GOL.store_name}账单明细{date_str}.xlsx")
        # 其他文件地址
        GOL.save_path.operate_detail = os.path.join(save_folder, f"{GOL.store_name}营业明细表{date_str}.xlsx",)
        GOL.save_path.synthesize_operate = os.path.join(save_folder, f"{GOL.store_name}综合营业统计{date_str}.xlsx")
        GOL.save_path.synthesize_income = os.path.join(save_folder, f"{GOL.store_name}综合收款统计{date_str}.xlsx")
        GOL.save_path.store_consume = os.path.join(save_folder, f"{GOL.store_name}储值消费汇总表{date_str}.xlsx")
        GOL.save_path.member_addition = os.path.join(save_folder, f"{GOL.store_name}会员新增情况统计表{date_str}.xlsx")
        GOL.save_path.pay_settlement = os.path.join(save_folder, f"{GOL.store_name}支付结算{date_str}.xlsx")
        GOL.save_path.pay_detail = os.path.join(save_folder, f"{GOL.store_name}支付明细{date_str}.xlsx")
        GOL.save_path.autotrophy_meituan = os.path.join(save_folder, f"{GOL.store_name}自营外卖{date_str}(美团后台).xlsx")
        GOL.save_path.autotrophy_dada = os.path.join(save_folder, f"{GOL.store_name}自营外卖{date_str}(达达).xlsx")
        if os.path.exists(GOL.save_path.operate_detail):
            print(f"汇总表格已在不跑第二遍:{GOL.save_path.operate_detail}")
            continue
        # 从网站上下载相关EXCEL文件
        crawler_main(chrome_path, chrome_driver_path, download_path, user_path)
        # 备份所有数据，作为回退处理，减少二次从网站上下载数据的情况发生
        copy_folder(save_folder, backup_folder)
        # 饿了么导出数据的处理-账单明细表
        eleme_main()
        # 自营外卖(美团后台)表处理
        meituan_autotrophy_main()
        # 自营外卖(达达)表处理
        dada_autotrophy_main()
        # 汇总外卖收入表
        take_out_main()
        # 汇总营业明细表
        operation_detail_main(operate_detail_template)

if __name__ == "__main__":
    main()
