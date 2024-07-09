import xlwt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from xlwt.Worksheet import Worksheet
from itertools import combinations
from typing import List, Dict, Tuple
from difflib import SequenceMatcher


class Unit:
    """每个单元类"""

    def __init__(self) -> None:
        self.index: List[int] = []
        self.amount: List[int] = []


def similarity_match(a, b):
    """定义相似度匹配函数"""
    similarity = SequenceMatcher(None, a, b).ratio()
    return similarity >= 0.8


def get_database(path):
    """"获取各个终端客户的产品名称转化为标准名称的数据库"""
    rd = {}
    df = pd.read_excel(path, header=1)
    for _, row in df.iterrows():
        rd[row["合并项"]] = row["品规(清洗后)"]
    return rd


def get_zgzy_data(path):
    """获取中国中药的数据及客户名称转化为标注名称的数据库"""
    rd: Dict[str, Dict[str, Unit]] = {}
    name2standard: Dict[str, Dict[str, str]] = {}
    raw_df = pd.read_excel(path, header=0)

    df = raw_df[["销售日期", "购入客户名称(原始)", "购入客户名称(清洗后)", "品规(清洗后)", "标准批号", "数量"]]
    for index, row in df.iterrows():
        name = row["购入客户名称(原始)"]
        name_standard = row["购入客户名称(清洗后)"]
        amount = row["数量"]
        key = "@@".join([row["销售日期"], row["品规(清洗后)"], row["标准批号"]])
        if key not in name2standard:
            name2standard[key] = {}
        name2standard[key][name] = name_standard

        if key not in rd:
            rd[key] = {name_standard: Unit()}
        if name_standard not in rd[key]:
            rd[key][name_standard] = Unit()
        rd[key][name_standard].amount.append(amount)
        rd[key][name_standard].index.append(index)

    return_df = raw_df[["销售日期", "购入客户名称(原始)", "品规(清洗后)", "标准批号", "数量"]]
    return_df = return_df.rename(columns={
        "购入客户名称(原始)": "客户名称",
        "标准批号": "批号",
        "数量": "销售数量"
    })
    return_df = return_df.reindex(columns=["销售日期", "客户名称", "品规(清洗后)", "批号", "销售数量"])
    return_df["来源"] = ["中国中药表"] * len(return_df)
    return return_df, rd, name2standard


def get_client_data(path, client_database: Dict[str, Dict[str, str]]
                    ) -> Tuple[pd.DataFrame, Dict[str, Dict[str, Unit]]]:
    """整理终端客户的数据"""
    rd: Dict[str, Dict[str, Unit]] = {}
    raw_df = pd.read_excel(path, header=0)
    df = raw_df[["销售日期", "客户名称", "品规(清洗后)", "批号", "销售数量"]]
    for index, row in df.iterrows():
        quality_regulation = row["品规(清洗后)"]
        date: pd.Timestamp = row["销售日期"]

        key = "@@".join([date.strftime("%Y-%m-%d"), quality_regulation, str(row["批号"])])

        name_standard = row["客户名称"]
        if key not in client_database:
            print(f"[警告]客户名匹配异常,日期-品类-批号未找到数据，请核实:{key}, {name_standard}")
        elif name_standard in client_database[key]:
            name_standard = client_database[key][name_standard]
        else:
            similarity_list = []
            for name in client_database[key].values():
                if not similarity_match(name, name_standard):
                    continue
                similarity_list.append(name)
            if len(similarity_list) != 1:
                print(f"[警告]客户名匹配异常, 未找到匹配项，请核实:{key}, {name_standard}, {similarity_list}")
            else:
                name_standard = similarity_list[0]

        if key not in rd:
            rd[key] = {name_standard: Unit()}
        if name_standard not in rd[key]:
            rd[key][name_standard] = Unit()
        rd[key][name_standard].amount.append(row["销售数量"])
        rd[key][name_standard].index.append(index)

    return_df = raw_df.reindex(columns=["销售日期", "客户名称", "品规(清洗后)", "批号", "销售数量"])
    return_df["来源"] = ["客户表"] * len(return_df)
    return return_df, rd


def find_combination(nums, target):
    """找到nums中所有可能的组合,其和等于target"""
    for r in range(1, len(nums) + 1):
        for combo in combinations(nums, r):
            if sum(combo) == target:
                return combo
    return None


def match_and_diff(list1: List, list2: List):
    """匹配不同的值"""
    list1 = list1.copy()
    list2 = list2.copy()
    unmatched_list1 = []
    unmatched_list2 = []

    while len(list1) != 0 and len(list2) != 0:
        max1 = max(list1)
        max2 = max(list2)
        if max1 == max2:
            list1.remove(max1)
            list2.remove(max2)
            continue
        if max1 > max2:
            list1.remove(max1)
            combo_for_max1 = find_combination(list2, max1)
            if combo_for_max1:
                for num in combo_for_max1:
                    list2.remove(num)
            else:
                unmatched_list1.append(max1)
        else:
            list2.remove(max2)
            combo_for_max2 = find_combination(list1, max2)
            if combo_for_max2:
                for num in combo_for_max2:
                    list1.remove(num)
            else:
                unmatched_list2.append(max2)
    unmatched_list1.extend(list1)
    unmatched_list2.extend(list2)
    return unmatched_list1, unmatched_list2


def fill_color(path, indexs):
    """根据索引填充颜色"""
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    wb = load_workbook(path)
    ws = wb.active
    for i in indexs:
        ws.cell(row=i + 2, column=1).fill = red_fill
    wb.save(path)


def get_xlwt_color_style(color):
    """获取颜色样式"""
    style = xlwt.XFStyle()
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = xlwt.Style.colour_map[color]
    style.pattern = pattern
    return style


def main():
    zgzy_path = r"E:\NewFolder\liuxiang\中国中药表 - 副本.xlsx"
    client_path = r"E:\NewFolder\liuxiang\客户表 - 副本.xlsx"
    print("读取中国中药数据表")
    zgzy_df, zgzy_dict, client_database = get_zgzy_data(zgzy_path)
    print("读取终端客户的数据表")
    client_df, client_dict = get_client_data(client_path, client_database)
    print("开始比对数据，获取差异项")
    zgzy_different = []
    client_different = []
    for client_k1, client_v1 in client_dict.items():
        if client_k1 not in zgzy_dict:
            for _, client_v2 in client_v1.items():
                client_different.extend(client_v2.index)
            continue
        zgzy_v1 = zgzy_dict[client_k1]
        for client_k2, client_v2 in client_v1.items():
            if client_k2 not in zgzy_v1:
                client_different.extend(client_v2.index)
                continue
            zgzy_v2 = zgzy_v1.pop(client_k2)
            unmatched_zgzy, unmatched_client = match_and_diff(zgzy_v2.amount, client_v2.amount)
            for amount in unmatched_zgzy:
                amount_index = zgzy_v2.amount.index(amount)
                zgzy_v2.amount.pop(amount_index)
                zgzy_different.append(zgzy_v2.index.pop(amount_index))
            for amount in unmatched_client:
                amount_index = client_v2.amount.index(amount)
                client_v2.amount.pop(amount_index)
                client_different.append(client_v2.index.pop(amount_index))
    # 补充中国中药里面的差异数据
    for _, v1 in zgzy_dict.items():
        for _, v2 in v1.items():
            zgzy_different.extend(v2.index)
    print("针对差异数据,在原始表上进行标红")
    fill_color(zgzy_path, zgzy_different)
    fill_color(client_path, client_different)
    print("输出差异表")
    yellow = get_xlwt_color_style("yellow")
    orange = get_xlwt_color_style("orange")
    wb = xlwt.Workbook()
    ws: Worksheet = wb.add_sheet("Sheet1")
    for j, value in enumerate(zgzy_df.columns):
        ws.write(0, j, value)
    row_i = 0
    for index, row in zgzy_df.iterrows():
        row_i += 1
        for j, value in enumerate(row):
            if index in zgzy_different:
                ws.write(row_i, j, value, yellow)
            else:
                ws.write(row_i, j, value)
    client_df["销售日期"] = client_df['销售日期'] = client_df['销售日期'].dt.strftime('%Y-%m-%d')
    for index, row in client_df.iterrows():
        row_i += 1
        for j, value in enumerate(row):
            if index in client_different:
                ws.write(row_i, j, value, orange)
            else:
                ws.write(row_i, j, value)
    wb.save(r"E:\NewFolder\liuxiang\核查报告.xls")
    print("程序已运行完毕")


if __name__ == '__main__':
    main()
